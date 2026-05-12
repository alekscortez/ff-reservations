#!/usr/bin/env python3
"""Extract + dedupe customer contacts from the legacy reservations.xlsx workbook.

The workbook has ~90 sheets, one per event date. Each data sheet has a header
row with columns: TABLE | NAME | PHONE NUMBER | SECTION | PRICE | STATUS | NOTES |
RESERVATION ID | QR | Zelle | Image. We collect every row with a (name, phone),
normalize phones to E.164, deduplicate by normalized phone, and aggregate:
  - bestName (most frequent non-placeholder spelling, ties → longest)
  - totalReservations (count of rows)
  - totalSpend (sum of price for paid statuses)
  - firstEventDate / lastEventDate (ISO dates)
  - alternateNames (other spellings seen)

Writes:
  - <out_dir>/contacts_raw.json     — every row that had a (name, phone)
  - <out_dir>/contacts_dedup.json   — deduped, one record per normalized phone
  - <out_dir>/contacts_dedup.csv    — same as above, CSV form for review/import
  - <out_dir>/contacts_dropped.json — rows dropped (invalid phone or placeholder)
  - <out_dir>/contacts_summary.json — top-level counts
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


# ---------- helpers ----------

PLACEHOLDER_NAME_PATTERNS = [
    re.compile(r"^\s*\*+\s*SPARE.*\*+\s*$", re.IGNORECASE),
    re.compile(r"^\s*SPARE\s*TABLE\s*$", re.IGNORECASE),
    re.compile(r"^\s*RESERVED\s*$", re.IGNORECASE),
    re.compile(r"^\s*N/?A\s*$", re.IGNORECASE),
    re.compile(r"^\s*-+\s*$"),
    re.compile(r"^\s*\?+\s*$"),
]

# Statuses we treat as "paid" → contributes to totalSpend.
# Empty or 'Not yet paid' → reservation not realized → don't count totalSpend
# 'Free' / 'Promotion' → reservation realized but no money
PAID_STATUSES = {
    "cash app", "cashapp", "cash", "square", "zelle",
    "card", "credit", "credit card", "debit", "venmo", "paypal", "transfer",
    "paid",
}
NON_PAID_STATUSES = {
    "not yet paid", "no pago", "free", "promotion", "promo", "comp", "courtesy",
    "other", "spare", "n/a", "",
}


def normalize_phone_digits(value) -> str:
    """Return digits-only string from any input. Floats like 9563108746.0 → '9563108746'."""
    if value is None:
        return ""
    if isinstance(value, float):
        # excel often stores 10-digit phones as floats; strip trailing .0
        if value.is_integer():
            value = str(int(value))
        else:
            value = str(value)
    elif isinstance(value, int):
        value = str(value)
    else:
        value = str(value)
    return re.sub(r"\D", "", value)


def to_e164(digits: str, default_country: str = "US") -> Optional[tuple[str, str]]:
    """Map a digit string to (e164, country). Returns None if unparseable.

    Heuristics tuned for this dataset (NAFTA / TX-MX border):
      - 10 digits → assume default_country (US in this dataset; 956 area code)
      - 11 digits starting with '1' → +1 US/CA
      - 11 digits starting with '52' → drop leading 5? no — '52' alone is 2 digits.
        12 digits starting with '52' → +52 MX
      - 13+ digits starting with '521' → +52 MX (mobile prefix), strip the '1'
    """
    if not digits:
        return None
    n = len(digits)
    if n == 10:
        return (f"+1{digits}", default_country)
    if n == 11 and digits.startswith("1"):
        return (f"+{digits}", "US")
    if n == 12 and digits.startswith("52"):
        return (f"+{digits}", "MX")
    if n == 13 and digits.startswith("521"):
        # MX mobile legacy prefix; canonicalize to +52 + last 10
        return (f"+52{digits[-10:]}", "MX")
    if n == 11 and digits.startswith("52"):
        # weird 11-digit MX; pad with assumption
        return None
    if n < 10:
        return None
    # >=14 digits, likely concatenated trash
    if n > 13:
        # Try the last 10 as US fallback
        last10 = digits[-10:]
        return (f"+1{last10}", default_country)
    return None


def is_placeholder_name(name: str) -> bool:
    if not name:
        return True
    s = name.strip()
    if not s:
        return True
    for pat in PLACEHOLDER_NAME_PATTERNS:
        if pat.match(s):
            return True
    return False


def clean_name(name) -> str:
    if name is None:
        return ""
    s = str(name).strip()
    # collapse internal whitespace
    s = re.sub(r"\s+", " ", s)
    return s


def status_is_paid(status) -> bool:
    if status is None:
        return False
    s = str(status).strip().lower()
    if s in PAID_STATUSES:
        return True
    if s in NON_PAID_STATUSES:
        return False
    # unknown / other → conservative: don't count as paid
    return False


def parse_price(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


SHEET_NAME_DATE_PATTERNS = [
    # 11-2-2024, 11-15-2025, 01-17-2026
    (re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{4})"), ("m", "d", "y4")),
    # 11-2-25, 11-9-24, 12-7-24
    (re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{2})$"), ("m", "d", "y2")),
    # 8302025 (mmddyyyy with no separators)
    (re.compile(r"^(\d{2})(\d{2})(\d{4})$"), ("m", "d", "y4")),
]


def parse_sheet_name_date(name: str) -> Optional[dt.date]:
    s = name.strip()
    for pat, parts in SHEET_NAME_DATE_PATTERNS:
        m = pat.match(s)
        if not m:
            continue
        try:
            mm = int(m.group(1))
            dd = int(m.group(2))
            yy = m.group(3)
            year = int(yy) if len(yy) == 4 else 2000 + int(yy)
            return dt.date(year, mm, dd)
        except (ValueError, IndexError):
            continue
    return None


def detect_event_date(ws) -> Optional[dt.date]:
    """Look at the first ~15 rows for a 'DATE' label whose neighbour cell holds a date."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 15:
            break
        if not row:
            continue
        for j, val in enumerate(row):
            if isinstance(val, str) and val.strip().upper() == "DATE":
                # look right for a date value
                for k in range(j + 1, min(j + 5, len(row))):
                    cand = row[k]
                    if isinstance(cand, dt.datetime):
                        return cand.date()
                    if isinstance(cand, dt.date):
                        return cand
                    if isinstance(cand, str):
                        try:
                            return dt.datetime.fromisoformat(cand).date()
                        except ValueError:
                            pass
    return None


def find_header_row(ws, max_rows: int = 20) -> Optional[tuple[int, dict[str, int]]]:
    """Return (row_index_0based, {column_name_upper: col_index_0based}) for the header row.

    Header row is identified by containing both 'NAME' and 'PHONE NUMBER' (or similar)
    in the same row.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > max_rows:
            return None
        if not row:
            continue
        cols = {}
        for j, val in enumerate(row):
            if isinstance(val, str):
                key = val.strip().upper()
                if key:
                    cols[key] = j
        # need at least NAME + PHONE NUMBER
        has_name = "NAME" in cols
        has_phone = (
            "PHONE NUMBER" in cols or "PHONE" in cols or "TELEFONO" in cols or "TELÉFONO" in cols
        )
        if has_name and has_phone:
            return (i, cols)
    return None


def col_or(cols: dict[str, int], *keys: str) -> Optional[int]:
    for k in keys:
        if k in cols:
            return cols[k]
    return None


# ---------- core extraction ----------

def extract_workbook(workbook_path: Path) -> dict:
    wb = openpyxl.load_workbook(str(workbook_path), read_only=False, data_only=True)
    raw_rows: list[dict] = []
    dropped: list[dict] = []
    skipped_sheets: list[str] = []
    parsed_sheets: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = find_header_row(ws)
        if not header:
            skipped_sheets.append(sheet_name)
            continue
        header_row_idx, cols = header

        col_table = col_or(cols, "TABLE", "MESA")
        col_name = col_or(cols, "NAME", "NOMBRE")
        col_phone = col_or(cols, "PHONE NUMBER", "PHONE", "TELEFONO", "TELÉFONO")
        col_section = col_or(cols, "SECTION", "SECCION", "SECCIÓN")
        col_price = col_or(cols, "PRICE", "PRECIO")
        col_status = col_or(cols, "STATUS", "ESTATUS", "ESTADO")
        col_notes = col_or(cols, "NOTES", "NOTAS")
        col_resid = col_or(cols, "RESERVATION ID", "RESERVATION", "ID")

        if col_name is None or col_phone is None:
            skipped_sheets.append(sheet_name)
            continue

        # detect event date
        ws_date = detect_event_date(ws) or parse_sheet_name_date(sheet_name)
        date_iso = ws_date.isoformat() if ws_date else None

        sheet_rows = 0
        sheet_with_phone = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_row_idx:
                continue
            if not row:
                continue

            def cell(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            raw_name = cell(col_name)
            raw_phone = cell(col_phone)
            raw_status = cell(col_status)
            raw_price = cell(col_price)
            raw_section = cell(col_section)
            raw_notes = cell(col_notes)
            raw_table = cell(col_table)
            raw_resid = cell(col_resid)

            name_str = clean_name(raw_name)
            phone_digits = normalize_phone_digits(raw_phone)

            # skip empty rows
            if not name_str and not phone_digits:
                continue

            sheet_rows += 1
            row_record = {
                "sheet": sheet_name,
                "rowIndex": i,
                "eventDate": date_iso,
                "table": raw_table,
                "name": name_str,
                "phoneRaw": str(raw_phone) if raw_phone is not None else "",
                "phoneDigits": phone_digits,
                "section": raw_section,
                "price": parse_price(raw_price),
                "status": (str(raw_status).strip() if raw_status is not None else ""),
                "notes": (str(raw_notes).strip() if raw_notes is not None else ""),
                "reservationId": (str(raw_resid).strip() if raw_resid is not None else ""),
            }

            if not phone_digits:
                # capture for the dropped report (no phone → can't dedupe)
                dropped.append({**row_record, "dropReason": "no_phone"})
                continue

            sheet_with_phone += 1

            if is_placeholder_name(name_str):
                dropped.append({**row_record, "dropReason": "placeholder_name"})
                continue

            e164 = to_e164(phone_digits)
            if not e164:
                dropped.append({**row_record, "dropReason": f"unparseable_phone_len_{len(phone_digits)}"})
                continue

            row_record["phoneE164"], row_record["phoneCountry"] = e164
            raw_rows.append(row_record)

        parsed_sheets.append({
            "sheet": sheet_name,
            "eventDate": date_iso,
            "rows": sheet_rows,
            "rowsWithPhone": sheet_with_phone,
        })

    return {
        "raw_rows": raw_rows,
        "dropped": dropped,
        "skipped_sheets": skipped_sheets,
        "parsed_sheets": parsed_sheets,
    }


# ---------- dedupe ----------

def dedupe_contacts(raw_rows: list[dict]) -> list[dict]:
    by_phone: dict[str, list[dict]] = {}
    for r in raw_rows:
        by_phone.setdefault(r["phoneE164"], []).append(r)

    deduped: list[dict] = []
    for phone, rows in by_phone.items():
        # pick best name: most frequent (case-insensitive), tie-break by length (longer wins),
        # tie-break by preferring titlecase / non-ALLCAPS
        name_counts: Counter[str] = Counter()
        for r in rows:
            if r["name"]:
                name_counts[r["name"]] += 1
        # group case-insensitively
        ci_groups: dict[str, list[tuple[str, int]]] = {}
        for nm, cnt in name_counts.items():
            ci_groups.setdefault(nm.lower(), []).append((nm, cnt))

        def variant_key(item: tuple[str, int]) -> tuple:
            nm, cnt = item
            # prefer not-all-caps spelling
            allcaps_penalty = 1 if nm.isupper() else 0
            return (-cnt, allcaps_penalty, -len(nm), nm)

        best_per_group = []
        for group, variants in ci_groups.items():
            variants_sorted = sorted(variants, key=variant_key)
            top_variant, top_cnt = variants_sorted[0]
            total_cnt = sum(c for _, c in variants)
            best_per_group.append((top_variant, total_cnt))
        best_per_group.sort(key=lambda x: (-x[1], -len(x[0]), x[0]))
        best_name = best_per_group[0][0] if best_per_group else ""

        alternate_names = [nm for nm, _ in best_per_group[1:5]]

        # totals
        total_reservations = len(rows)
        total_spend = 0.0
        for r in rows:
            if status_is_paid(r["status"]) and r["price"] is not None:
                total_spend += r["price"]

        dates = sorted([r["eventDate"] for r in rows if r["eventDate"]])
        first_event = dates[0] if dates else None
        last_event = dates[-1] if dates else None

        # phoneCountry: most-common across rows
        country_counts = Counter(r["phoneCountry"] for r in rows if r.get("phoneCountry"))
        phone_country = country_counts.most_common(1)[0][0] if country_counts else "US"

        # any notes seen? (concat unique)
        notes_seen = sorted({r["notes"] for r in rows if r.get("notes")})

        deduped.append({
            "phoneE164": phone,
            "phoneCountry": phone_country,
            "name": best_name,
            "alternateNames": alternate_names,
            "totalReservations": total_reservations,
            "totalSpend": round(total_spend, 2),
            "firstEventDate": first_event,
            "lastEventDate": last_event,
            "notes": " | ".join(notes_seen) if notes_seen else "",
        })

    # sort by totalReservations desc, then totalSpend desc, then name
    deduped.sort(key=lambda c: (-c["totalReservations"], -c["totalSpend"], c["name"].lower()))
    return deduped


# ---------- main ----------

def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--xlsx",
        default="/Users/alekscortez/WebstormProjects/ff-reservations/src/assets/reservations.xlsx",
    )
    p.add_argument(
        "--out-dir",
        default="/Users/alekscortez/WebstormProjects/ff-reservations/scripts/out",
    )
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"error: workbook not found: {xlsx_path}", file=sys.stderr)
        return 1

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"reading {xlsx_path}…")
    try:
        result = extract_workbook(xlsx_path)
    except InvalidFileException as e:
        print(f"error: failed to read workbook: {e}", file=sys.stderr)
        return 2

    raw = result["raw_rows"]
    dropped = result["dropped"]
    skipped = result["skipped_sheets"]
    parsed = result["parsed_sheets"]

    print(f"parsed sheets: {len(parsed)}")
    print(f"skipped sheets (no NAME/PHONE header): {len(skipped)}")
    if skipped:
        print("  →", ", ".join(skipped))
    print(f"raw contact rows (with E.164 phone + non-placeholder name): {len(raw)}")
    print(f"dropped rows: {len(dropped)}")
    drop_reasons = Counter(d["dropReason"] for d in dropped)
    for reason, cnt in drop_reasons.most_common():
        print(f"  - {reason}: {cnt}")

    deduped = dedupe_contacts(raw)
    print(f"unique contacts after dedupe: {len(deduped)}")

    # write outputs
    (out_dir / "contacts_raw.json").write_text(json.dumps(raw, indent=2, default=str))
    (out_dir / "contacts_dedup.json").write_text(json.dumps(deduped, indent=2, default=str))
    (out_dir / "contacts_dropped.json").write_text(json.dumps(dropped, indent=2, default=str))

    summary = {
        "workbook": str(xlsx_path),
        "parsedSheetCount": len(parsed),
        "skippedSheetCount": len(skipped),
        "skippedSheets": skipped,
        "rawRowCount": len(raw),
        "droppedRowCount": len(dropped),
        "dropReasons": dict(drop_reasons),
        "uniqueContactCount": len(deduped),
        "perSheet": parsed,
    }
    (out_dir / "contacts_summary.json").write_text(json.dumps(summary, indent=2, default=str))

    csv_path = out_dir / "contacts_dedup.csv"
    with csv_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "phoneE164", "phoneCountry", "name", "alternateNames",
            "totalReservations", "totalSpend", "firstEventDate", "lastEventDate", "notes",
        ])
        for c in deduped:
            w.writerow([
                c["phoneE164"], c["phoneCountry"], c["name"],
                "; ".join(c["alternateNames"]),
                c["totalReservations"], c["totalSpend"],
                c["firstEventDate"] or "", c["lastEventDate"] or "",
                c["notes"],
            ])

    print(f"\nwrote:")
    for f in ["contacts_raw.json", "contacts_dedup.json", "contacts_dropped.json", "contacts_summary.json", "contacts_dedup.csv"]:
        print(f"  {out_dir / f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
